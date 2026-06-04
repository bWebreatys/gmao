"""
GMAO – Application de saisie locale complète
Onglets : EQUIPEMENTS, COMPOSANTS, PACKAGES, MAINT_PREV,
          INTERVENTIONS, FOURNISSEURS, PIECES, TECHNICIENS
Usage : python app.py  →  http://127.0.0.1:5000
"""

import re, threading, webbrowser
from datetime import datetime, date
from pathlib import Path
from flask import (Flask, render_template, request,
                   redirect, url_for, flash, jsonify)
from openpyxl import load_workbook

XLSX_PATH = Path(__file__).parent.parent / "GMAO_v0.1.xlsx"
app = Flask(__name__)
app.secret_key = "gmao-local-2025"

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def load_wb(read_only=False):
    return load_workbook(str(XLSX_PATH), read_only=read_only, data_only=False)

def get_listes():
    wb = load_wb()
    ws = wb["LISTES"]
    out = {}
    for col in ws.iter_cols():
        cells = list(col)
        key = cells[0].value
        vals = [c.value for c in cells[1:] if c.value is not None]
        if key:
            out[key] = vals
    wb.close()
    return out

def get_col_values(sheet_name, col_index):
    wb = load_wb(read_only=True)
    ws = wb[sheet_name]
    vals = []
    for row in ws.iter_rows(min_row=2, min_col=col_index,
                             max_col=col_index, values_only=True):
        v = row[0]
        if v is not None:
            sv = str(v)
            if not sv.startswith("="):
                vals.append(sv)
    wb.close()
    return vals

def parse_date_fr(s):
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def fmt_date(val):
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    d = parse_date_fr(str(val))
    return d.strftime("%d/%m/%Y") if d else str(val)

def is_formula(v):
    return isinstance(v, str) and v.startswith("=")

def next_code(prefix, existing, digits=4):
    nums = []
    pat = re.compile(rf"^{re.escape(prefix)}(\d+)$", re.IGNORECASE)
    for c in existing:
        m = pat.match(str(c))
        if m:
            nums.append(int(m.group(1)))
    return f"{prefix}{(max(nums, default=0)+1):0{digits}d}"

def validate_dates(data, date_fields, errors):
    for f in date_fields:
        v = data.get(f, "").strip()
        if v and not parse_date_fr(v):
            errors[f] = "Format attendu : JJ/MM/AAAA"

def read_sheet(sheet_name, columns):
    wb = load_wb(read_only=True)
    ws = wb[sheet_name]
    rows = []
    for r in ws.iter_rows(min_row=2):
        if r[0].value is None:
            continue
        row = {}
        for col_idx, label, typ in columns:
            val = r[col_idx - 1].value
            if is_formula(val):
                row[label] = ""
            elif typ == "date":
                row[label] = fmt_date(val)
            else:
                row[label] = val if val is not None else ""
        rows.append(row)
    wb.close()
    return rows

def read_one(sheet_name, columns, pk_value):
    for r in read_sheet(sheet_name, columns):
        if str(r.get(columns[0][1], "")) == str(pk_value):
            return r
    return None

def find_row_num(ws, pk_value):
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(pk_value):
            return i
    return None

def write_row(sheet_name, columns, data, is_new, formula_cols=None):
    if formula_cols is None:
        formula_cols = set()
    wb = load_wb()
    ws = wb[sheet_name]
    if is_new:
        row_num = None
        for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
            if r[0].value is None:
                row_num = i
                break
        if row_num is None:
            row_num = ws.max_row + 1
    else:
        row_num = find_row_num(ws, data[columns[0][1]])
        if row_num is None:
            return False, "Enregistrement introuvable."
    for col_idx, label, typ in columns:
        if col_idx in formula_cols:
            continue
        cell = ws.cell(row=row_num, column=col_idx)
        if is_formula(cell.value):
            continue
        val = data.get(label, "")
        if typ == "date":
            d = parse_date_fr(str(val)) if val else None
            cell.value = d
            if d:
                cell.number_format = "DD/MM/YYYY"
        elif typ == "float":
            try:
                cell.value = float(val) if val not in (None, "") else None
            except (ValueError, TypeError):
                cell.value = str(val) if val else None
        elif typ == "int":
            try:
                cell.value = int(val) if val not in (None, "") else None
            except (ValueError, TypeError):
                cell.value = str(val) if val else None
        else:
            cell.value = val if val not in (None, "") else None
    wb.save(str(XLSX_PATH))
    wb.close()
    return True, "Enregistré avec succès."

def delete_row(sheet_name, columns, pk_value):
    wb = load_wb()
    ws = wb[sheet_name]
    row_num = find_row_num(ws, pk_value)
    if row_num is None:
        return False, "Enregistrement introuvable."
    for col_idx, label, typ in columns:
        cell = ws.cell(row=row_num, column=col_idx)
        if not is_formula(cell.value):
            cell.value = None
    wb.save(str(XLSX_PATH))
    wb.close()
    return True, "Supprimé avec succès."

# ══════════════════════════════════════════════════════════════════════════════
# DÉFINITIONS DES COLONNES
# ══════════════════════════════════════════════════════════════════════════════

EQ_COLS = [
    (1,"Code_Equipement","text"),(2,"Designation","text"),(3,"Categorie","select"),
    (4,"Criticite","select"),(5,"Service","select"),(6,"Localisation","text"),
    (7,"Marque","text"),(8,"Modele","text"),(9,"N_Serie","text"),
    (10,"Date_Acquisition","date"),(11,"Fournisseur","fk"),(12,"Garantie_Fin","date"),
    (13,"Etat","select"),(14,"Date_Derniere_Maintenance","date"),
    (15,"Prochaine_Maintenance","date"),(16,"Periodicite_Mois","select"),
    (17,"Technicien_Referent","fk"),(18,"Code_Package","fk"),
    (19,"Nb_Composants","formula"),(20,"Composant_Defectueux","formula"),
    (21,"Observation","textarea"),
    (22,"Nom_Complementaire","text"),(23,"Provenance","select"),
    (24,"Ile","select"),(25,"Conformite_Charte_Dons","select"),
    (26,"Reforme","select"),(27,"Date_Reforme","date"),
]
EQ_FCOLS = {19,20}
EQ_DATES = ["Date_Acquisition","Garantie_Fin","Date_Derniere_Maintenance","Prochaine_Maintenance","Date_Reforme"]

COMP_COLS = [
    (1,"Code_Composant","text"),(2,"Code_Equipement","fk"),(3,"Designation_Equip","formula"),
    (4,"Nom_Composant","text"),(5,"Type_Composant","select"),(6,"Marque_Composant","text"),
    (7,"Statut_Composant","select"),(8,"Date_Detection_Defaut","date"),
    (9,"Action_Requise","select"),(10,"Date_Commande","date"),
    (11,"Fournisseur_Piece","fk"),(12,"Ref_Piece_Commandee","text"),
    (13,"Date_Remplacement","date"),(14,"Technicien","fk"),(15,"Code_Package","fk"),
    (16,"Observation","textarea"),
]
COMP_FCOLS = {3}
COMP_DATES = ["Date_Detection_Defaut","Date_Commande","Date_Remplacement"]
COMP_ACTIONS = ["Aucune","Remplacement","Commande en cours","Réparation sur site","Retour fabricant"]

PKG_COLS = [
    (1,"Code_Package","text"),(2,"Nom_Package","text"),(3,"Description","textarea"),
    (4,"Service","select"),(5,"Technicien_Referent","fk"),(6,"Date_Creation","date"),
    (7,"Statut_Package","select"),(8,"Nb_Equip_Package","formula"),
    (9,"Nb_Composants_Def","formula"),(10,"Observation","textarea"),
]
PKG_FCOLS = {8,9}
PKG_DATES = ["Date_Creation"]
PKG_STATUTS = ["Actif","Incomplet","Hors service","Archivé"]

MP_COLS = [
    (1,"Code_MP","text"),(2,"Code_Equipement","fk"),(3,"Designation","formula"),
    (4,"Type_Maintenance","select"),(5,"Date_Prevue","date"),
    (6,"Technicien_Assigne","fk"),(7,"Statut","select"),
    (8,"Date_Realisation","date"),(9,"Temps_Passe_Heures","float"),
    (10,"Retard_Jours","formula"),(11,"Observation","textarea"),
]
MP_FCOLS = {3,10}
MP_DATES = ["Date_Prevue","Date_Realisation"]

INT_COLS = [
    (1,"Code_Intervention","text"),(2,"Code_Equipement","fk"),(3,"Designation","formula"),
    (4,"Type_Intervention","select"),(5,"Origine_Demande","select"),
    (6,"Date_Demande","date"),(7,"Date_Intervention","date"),
    (8,"Technicien","fk"),(9,"Description_Panne","textarea"),
    (10,"Action_Realisee","textarea"),(11,"Duree_Intervention_H","float"),
    (12,"Statut","select"),(13,"Pieces_Utilisees","text"),
    (14,"Code_Composant_Lie","fk"),(15,"Jours_Ouvert","formula"),
    (16,"Observation_Cloture","textarea"),
]
INT_FCOLS = {3,15}
INT_DATES = ["Date_Demande","Date_Intervention"]

FOUR_COLS = [
    (1,"Code_Fournisseur","text"),(2,"Nom_Fournisseur","text"),(3,"Type","select"),
    (4,"Contact","text"),(5,"Telephone","text"),(6,"Email","text"),
    (7,"Contrat","select"),(8,"Date_Debut_Contrat","date"),(9,"Date_Fin_Contrat","date"),
    (10,"Delai_Livraison_Jours","int"),(11,"Observation","textarea"),
]
FOUR_FCOLS = set()
FOUR_DATES = ["Date_Debut_Contrat","Date_Fin_Contrat"]

PIECE_COLS = [
    (1,"Code_Piece","text"),(2,"Designation","text"),(3,"Categorie","select"),
    (4,"Equipement_Compatible","fk"),(5,"Stock_Actuel","int"),(6,"Stock_Min","int"),
    (7,"Stock_Max","int"),(8,"Fournisseur","fk"),(9,"Prix_Unitaire","float"),
    (10,"Derniere_Reception","date"),(11,"Alerte_Stock","formula"),(12,"Observation","textarea"),
]
PIECE_FCOLS = {11}
PIECE_DATES = ["Derniere_Reception"]

TECH_COLS = [
    (1,"Code_Tech","text"),(2,"Nom","text"),(3,"Prenom","text"),
    (4,"Fonction","text"),(5,"Niveau_Qualification","select"),
    (6,"Experience_Annees","int"),(7,"Service_Rattachement","select"),
    (8,"Tel","text"),(9,"Email","text"),(10,"Formations","textarea"),
    (11,"Charge_Interventions","formula"),
]
TECH_FCOLS = {11}
TECH_DATES = []

# ══════════════════════════════════════════════════════════════════════════════
# ROUTES GÉNÉRIQUES (factory)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return redirect(url_for("equipements_list"))

@app.route("/api/suggest_code")
def api_suggest_code():
    prefix = request.args.get("prefix","EQ")
    sheet  = request.args.get("sheet","EQUIPEMENTS")
    digits = int(request.args.get("digits",4))
    return jsonify({"code": next_code(prefix, get_col_values(sheet,1), digits)})

# ── EQUIPEMENTS ───────────────────────────────────────────────────────────────

@app.route("/equipements")
def equipements_list():
    rows = read_sheet("EQUIPEMENTS", EQ_COLS)
    q = request.args.get("q","").strip().lower()
    fe = request.args.get("etat","").strip()
    fc = request.args.get("criticite","").strip()
    fi = request.args.get("ile","").strip()
    if q:  rows = [r for r in rows if any(q in str(v).lower() for v in r.values())]
    if fe: rows = [r for r in rows if r.get("Etat") == fe]
    if fc: rows = [r for r in rows if r.get("Criticite") == fc]
    if fi: rows = [r for r in rows if r.get("Ile") == fi]
    L = get_listes()
    return render_template("equipements_list.html", rows=rows,
        search=q, filtre_etat=fe, filtre_crit=fc, filtre_ile=fi, iles=L["Ile"],
        etats=L["Etat_Equip"], criticites=L["Criticite"])

@app.route("/equipements/nouveau", methods=["GET","POST"])
def equipements_new():
    L = get_listes()
    existing = get_col_values("EQUIPEMENTS",1)
    errors = {}
    fd = {lbl:"" for _,lbl,_ in EQ_COLS}
    fd["Code_Equipement"] = next_code("EQ", existing, 4)
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Code_Equipement","").strip(): errors["Code_Equipement"]="Obligatoire."
        elif fd["Code_Equipement"] in existing:      errors["Code_Equipement"]="Code déjà utilisé."
        if not fd.get("Designation","").strip(): errors["Designation"]="Obligatoire."
        if not fd.get("Etat","").strip():        errors["Etat"]="Obligatoire."
        if not fd.get("Criticite","").strip():   errors["Criticite"]="Obligatoire."
        if not fd.get("Service","").strip():     errors["Service"]="Obligatoire."
        if not fd.get("Categorie","").strip():   errors["Categorie"]="Obligatoire."
        validate_dates(fd, EQ_DATES, errors)
        if not errors:
            ok,msg = write_row("EQUIPEMENTS",EQ_COLS,fd,True,EQ_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("equipements_list"))
    return render_template("equipements_form.html", title="Nouvel équipement",
        action=url_for("equipements_new"), fd=fd, errors=errors, L=L,
        fournisseurs=get_col_values("FOURNISSEURS",1),
        techniciens=get_col_values("TECHNICIENS",1),
        packages=get_col_values("PACKAGES",1), is_new=True)

@app.route("/equipements/modifier/<code>", methods=["GET","POST"])
def equipements_edit(code):
    L = get_listes()
    errors = {}
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Designation","").strip(): errors["Designation"]="Obligatoire."
        if not fd.get("Etat","").strip():        errors["Etat"]="Obligatoire."
        if not fd.get("Criticite","").strip():   errors["Criticite"]="Obligatoire."
        validate_dates(fd, EQ_DATES, errors)
        if not errors:
            ok,msg = write_row("EQUIPEMENTS",EQ_COLS,fd,False,EQ_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("equipements_list"))
    else:
        fd = read_one("EQUIPEMENTS",EQ_COLS,code)
        if not fd: flash("Introuvable.","danger"); return redirect(url_for("equipements_list"))
    return render_template("equipements_form.html", title=f"Modifier équipement – {code}",
        action=url_for("equipements_edit",code=code), fd=fd, errors=errors, L=L,
        fournisseurs=get_col_values("FOURNISSEURS",1),
        techniciens=get_col_values("TECHNICIENS",1),
        packages=get_col_values("PACKAGES",1), is_new=False)

@app.route("/equipements/supprimer/<code>", methods=["POST"])
def equipements_delete(code):
    ok,msg = delete_row("EQUIPEMENTS",EQ_COLS,code)
    flash(msg,"success" if ok else "danger")
    return redirect(url_for("equipements_list"))

# ── COMPOSANTS ────────────────────────────────────────────────────────────────

@app.route("/composants")
def composants_list():
    rows = read_sheet("COMPOSANTS", COMP_COLS)
    q  = request.args.get("q","").strip().lower()
    fs = request.args.get("statut","").strip()
    fe = request.args.get("eq","").strip()
    if q:  rows = [r for r in rows if any(q in str(v).lower() for v in r.values())]
    if fs: rows = [r for r in rows if r.get("Statut_Composant") == fs]
    if fe: rows = [r for r in rows if r.get("Code_Equipement") == fe]
    L = get_listes()
    return render_template("composants_list.html", rows=rows,
        search=q, f_stat=fs, f_eq=fe,
        statuts=L["Statut_Composant"], equipements=get_col_values("EQUIPEMENTS",1))

@app.route("/composants/nouveau", methods=["GET","POST"])
def composants_new():
    L = get_listes()
    existing = get_col_values("COMPOSANTS",1)
    errors = {}
    fd = {lbl:"" for _,lbl,_ in COMP_COLS}
    fd["Code_Composant"] = next_code("COMP", existing, 3)
    if request.args.get("eq"): fd["Code_Equipement"] = request.args.get("eq")
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Code_Composant","").strip(): errors["Code_Composant"]="Obligatoire."
        elif fd["Code_Composant"] in existing:      errors["Code_Composant"]="Code déjà utilisé."
        if not fd.get("Code_Equipement","").strip():  errors["Code_Equipement"]="Obligatoire."
        if not fd.get("Nom_Composant","").strip():    errors["Nom_Composant"]="Obligatoire."
        if not fd.get("Statut_Composant","").strip(): errors["Statut_Composant"]="Obligatoire."
        validate_dates(fd, COMP_DATES, errors)
        if not errors:
            ok,msg = write_row("COMPOSANTS",COMP_COLS,fd,True,COMP_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("composants_list"))
    return render_template("composants_form.html", title="Nouveau composant",
        action=url_for("composants_new"), fd=fd, errors=errors, L=L,
        equipements=get_col_values("EQUIPEMENTS",1),
        fournisseurs=get_col_values("FOURNISSEURS",1),
        techniciens=get_col_values("TECHNICIENS",1),
        packages=get_col_values("PACKAGES",1),
        action_list=COMP_ACTIONS, is_new=True)

@app.route("/composants/modifier/<code>", methods=["GET","POST"])
def composants_edit(code):
    L = get_listes()
    errors = {}
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Nom_Composant","").strip():    errors["Nom_Composant"]="Obligatoire."
        if not fd.get("Statut_Composant","").strip(): errors["Statut_Composant"]="Obligatoire."
        validate_dates(fd, COMP_DATES, errors)
        if not errors:
            ok,msg = write_row("COMPOSANTS",COMP_COLS,fd,False,COMP_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("composants_list"))
    else:
        fd = read_one("COMPOSANTS",COMP_COLS,code)
        if not fd: flash("Introuvable.","danger"); return redirect(url_for("composants_list"))
    return render_template("composants_form.html", title=f"Modifier composant – {code}",
        action=url_for("composants_edit",code=code), fd=fd, errors=errors, L=L,
        equipements=get_col_values("EQUIPEMENTS",1),
        fournisseurs=get_col_values("FOURNISSEURS",1),
        techniciens=get_col_values("TECHNICIENS",1),
        packages=get_col_values("PACKAGES",1),
        action_list=COMP_ACTIONS, is_new=False)

@app.route("/composants/supprimer/<code>", methods=["POST"])
def composants_delete(code):
    ok,msg = delete_row("COMPOSANTS",COMP_COLS,code)
    flash(msg,"success" if ok else "danger")
    return redirect(url_for("composants_list"))

# ── PACKAGES ──────────────────────────────────────────────────────────────────

@app.route("/packages")
def packages_list():
    rows = read_sheet("PACKAGES", PKG_COLS)
    q  = request.args.get("q","").strip().lower()
    fs = request.args.get("statut","").strip()
    if q:  rows = [r for r in rows if any(q in str(v).lower() for v in r.values())]
    if fs: rows = [r for r in rows if r.get("Statut_Package") == fs]
    return render_template("packages_list.html", rows=rows,
        search=q, f_stat=fs, statuts=PKG_STATUTS)

@app.route("/packages/nouveau", methods=["GET","POST"])
def packages_new():
    L = get_listes()
    existing = get_col_values("PACKAGES",1)
    errors = {}
    fd = {lbl:"" for _,lbl,_ in PKG_COLS}
    fd["Code_Package"] = next_code("PKG", existing, 3)
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Code_Package","").strip(): errors["Code_Package"]="Obligatoire."
        elif fd["Code_Package"] in existing:      errors["Code_Package"]="Code déjà utilisé."
        if not fd.get("Nom_Package","").strip():  errors["Nom_Package"]="Obligatoire."
        validate_dates(fd, PKG_DATES, errors)
        if not errors:
            ok,msg = write_row("PACKAGES",PKG_COLS,fd,True,PKG_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("packages_list"))
    return render_template("packages_form.html", title="Nouveau package",
        action=url_for("packages_new"), fd=fd, errors=errors, L=L,
        techniciens=get_col_values("TECHNICIENS",1),
        statuts=PKG_STATUTS, is_new=True)

@app.route("/packages/modifier/<code>", methods=["GET","POST"])
def packages_edit(code):
    L = get_listes()
    errors = {}
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Nom_Package","").strip(): errors["Nom_Package"]="Obligatoire."
        validate_dates(fd, PKG_DATES, errors)
        if not errors:
            ok,msg = write_row("PACKAGES",PKG_COLS,fd,False,PKG_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("packages_list"))
    else:
        fd = read_one("PACKAGES",PKG_COLS,code)
        if not fd: flash("Introuvable.","danger"); return redirect(url_for("packages_list"))
    return render_template("packages_form.html", title=f"Modifier package – {code}",
        action=url_for("packages_edit",code=code), fd=fd, errors=errors, L=L,
        techniciens=get_col_values("TECHNICIENS",1),
        statuts=PKG_STATUTS, is_new=False)

@app.route("/packages/supprimer/<code>", methods=["POST"])
def packages_delete(code):
    ok,msg = delete_row("PACKAGES",PKG_COLS,code)
    flash(msg,"success" if ok else "danger")
    return redirect(url_for("packages_list"))

# ── MAINT_PREV ────────────────────────────────────────────────────────────────

@app.route("/maint_prev")
def maint_prev_list():
    rows = read_sheet("MAINT_PREV", MP_COLS)
    q  = request.args.get("q","").strip().lower()
    fs = request.args.get("statut","").strip()
    fe = request.args.get("eq","").strip()
    if q:  rows = [r for r in rows if any(q in str(v).lower() for v in r.values())]
    if fs: rows = [r for r in rows if r.get("Statut") == fs]
    if fe: rows = [r for r in rows if r.get("Code_Equipement") == fe]
    L = get_listes()
    return render_template("maint_prev_list.html", rows=rows,
        search=q, f_stat=fs, f_eq=fe,
        statuts=L["Statut_MP"], equipements=get_col_values("EQUIPEMENTS",1))

@app.route("/maint_prev/nouveau", methods=["GET","POST"])
def maint_prev_new():
    L = get_listes()
    existing = get_col_values("MAINT_PREV",1)
    errors = {}
    fd = {lbl:"" for _,lbl,_ in MP_COLS}
    fd["Code_MP"] = next_code("MP", existing, 4)
    fd["Statut"]  = "Planifiée"
    if request.args.get("eq"): fd["Code_Equipement"] = request.args.get("eq")
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Code_MP","").strip():          errors["Code_MP"]="Obligatoire."
        elif fd["Code_MP"] in existing:               errors["Code_MP"]="Code déjà utilisé."
        if not fd.get("Code_Equipement","").strip():  errors["Code_Equipement"]="Obligatoire."
        if not fd.get("Type_Maintenance","").strip(): errors["Type_Maintenance"]="Obligatoire."
        if not fd.get("Date_Prevue","").strip():      errors["Date_Prevue"]="Obligatoire."
        validate_dates(fd, MP_DATES, errors)
        if not errors:
            ok,msg = write_row("MAINT_PREV",MP_COLS,fd,True,MP_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("maint_prev_list"))
    return render_template("maint_prev_form.html", title="Nouvelle maintenance préventive",
        action=url_for("maint_prev_new"), fd=fd, errors=errors, L=L,
        equipements=get_col_values("EQUIPEMENTS",1),
        techniciens=get_col_values("TECHNICIENS",1), is_new=True)

@app.route("/maint_prev/modifier/<code>", methods=["GET","POST"])
def maint_prev_edit(code):
    L = get_listes()
    errors = {}
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Type_Maintenance","").strip(): errors["Type_Maintenance"]="Obligatoire."
        validate_dates(fd, MP_DATES, errors)
        if not errors:
            ok,msg = write_row("MAINT_PREV",MP_COLS,fd,False,MP_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("maint_prev_list"))
    else:
        fd = read_one("MAINT_PREV",MP_COLS,code)
        if not fd: flash("Introuvable.","danger"); return redirect(url_for("maint_prev_list"))
    return render_template("maint_prev_form.html", title=f"Modifier maintenance – {code}",
        action=url_for("maint_prev_edit",code=code), fd=fd, errors=errors, L=L,
        equipements=get_col_values("EQUIPEMENTS",1),
        techniciens=get_col_values("TECHNICIENS",1), is_new=False)

@app.route("/maint_prev/supprimer/<code>", methods=["POST"])
def maint_prev_delete(code):
    ok,msg = delete_row("MAINT_PREV",MP_COLS,code)
    flash(msg,"success" if ok else "danger")
    return redirect(url_for("maint_prev_list"))

# ── INTERVENTIONS ─────────────────────────────────────────────────────────────

@app.route("/interventions")
def interventions_list():
    rows = read_sheet("INTERVENTIONS", INT_COLS)
    q  = request.args.get("q","").strip().lower()
    fs = request.args.get("statut","").strip()
    ft = request.args.get("type","").strip()
    fe = request.args.get("eq","").strip()
    if q:  rows = [r for r in rows if any(q in str(v).lower() for v in r.values())]
    if fs: rows = [r for r in rows if r.get("Statut") == fs]
    if ft: rows = [r for r in rows if r.get("Type_Intervention") == ft]
    if fe: rows = [r for r in rows if r.get("Code_Equipement") == fe]
    L = get_listes()
    return render_template("interventions_list.html", rows=rows,
        search=q, f_stat=fs, f_type=ft, f_eq=fe,
        statuts=L["Statut_Intervention"], types=L["Type_Intervention"],
        equipements=get_col_values("EQUIPEMENTS",1))

@app.route("/interventions/nouveau", methods=["GET","POST"])
def interventions_new():
    L = get_listes()
    existing = get_col_values("INTERVENTIONS",1)
    errors = {}
    fd = {lbl:"" for _,lbl,_ in INT_COLS}
    fd["Code_Intervention"] = next_code("INT", existing, 4)
    fd["Statut"] = "Ouverte"
    if request.args.get("eq"): fd["Code_Equipement"] = request.args.get("eq")
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Code_Intervention","").strip(): errors["Code_Intervention"]="Obligatoire."
        elif fd["Code_Intervention"] in existing:      errors["Code_Intervention"]="Code déjà utilisé."
        if not fd.get("Code_Equipement","").strip():   errors["Code_Equipement"]="Obligatoire."
        if not fd.get("Type_Intervention","").strip():  errors["Type_Intervention"]="Obligatoire."
        if not fd.get("Date_Demande","").strip():       errors["Date_Demande"]="Obligatoire."
        validate_dates(fd, INT_DATES, errors)
        if not errors:
            ok,msg = write_row("INTERVENTIONS",INT_COLS,fd,True,INT_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("interventions_list"))
    return render_template("interventions_form.html", title="Nouvelle intervention",
        action=url_for("interventions_new"), fd=fd, errors=errors, L=L,
        equipements=get_col_values("EQUIPEMENTS",1),
        techniciens=get_col_values("TECHNICIENS",1),
        composants=get_col_values("COMPOSANTS",1), is_new=True)

@app.route("/interventions/modifier/<code>", methods=["GET","POST"])
def interventions_edit(code):
    L = get_listes()
    errors = {}
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Type_Intervention","").strip(): errors["Type_Intervention"]="Obligatoire."
        validate_dates(fd, INT_DATES, errors)
        if not errors:
            ok,msg = write_row("INTERVENTIONS",INT_COLS,fd,False,INT_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("interventions_list"))
    else:
        fd = read_one("INTERVENTIONS",INT_COLS,code)
        if not fd: flash("Introuvable.","danger"); return redirect(url_for("interventions_list"))
    return render_template("interventions_form.html", title=f"Modifier intervention – {code}",
        action=url_for("interventions_edit",code=code), fd=fd, errors=errors, L=L,
        equipements=get_col_values("EQUIPEMENTS",1),
        techniciens=get_col_values("TECHNICIENS",1),
        composants=get_col_values("COMPOSANTS",1), is_new=False)

@app.route("/interventions/supprimer/<code>", methods=["POST"])
def interventions_delete(code):
    ok,msg = delete_row("INTERVENTIONS",INT_COLS,code)
    flash(msg,"success" if ok else "danger")
    return redirect(url_for("interventions_list"))

# ── FOURNISSEURS ──────────────────────────────────────────────────────────────

@app.route("/fournisseurs")
def fournisseurs_list():
    rows = read_sheet("FOURNISSEURS", FOUR_COLS)
    q  = request.args.get("q","").strip().lower()
    ft = request.args.get("type","").strip()
    if q:  rows = [r for r in rows if any(q in str(v).lower() for v in r.values())]
    if ft: rows = [r for r in rows if r.get("Type") == ft]
    L = get_listes()
    return render_template("fournisseurs_list.html", rows=rows,
        search=q, f_type=ft, types=L["Type_Fournisseur"])

@app.route("/fournisseurs/nouveau", methods=["GET","POST"])
def fournisseurs_new():
    L = get_listes()
    existing = get_col_values("FOURNISSEURS",1)
    errors = {}
    fd = {lbl:"" for _,lbl,_ in FOUR_COLS}
    fd["Code_Fournisseur"] = next_code("FOUR", existing, 3)
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Code_Fournisseur","").strip(): errors["Code_Fournisseur"]="Obligatoire."
        elif fd["Code_Fournisseur"] in existing:      errors["Code_Fournisseur"]="Code déjà utilisé."
        if not fd.get("Nom_Fournisseur","").strip():  errors["Nom_Fournisseur"]="Obligatoire."
        validate_dates(fd, FOUR_DATES, errors)
        if not errors:
            ok,msg = write_row("FOURNISSEURS",FOUR_COLS,fd,True,FOUR_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("fournisseurs_list"))
    return render_template("fournisseurs_form.html", title="Nouveau fournisseur",
        action=url_for("fournisseurs_new"), fd=fd, errors=errors, L=L, is_new=True)

@app.route("/fournisseurs/modifier/<code>", methods=["GET","POST"])
def fournisseurs_edit(code):
    L = get_listes()
    errors = {}
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Nom_Fournisseur","").strip(): errors["Nom_Fournisseur"]="Obligatoire."
        validate_dates(fd, FOUR_DATES, errors)
        if not errors:
            ok,msg = write_row("FOURNISSEURS",FOUR_COLS,fd,False,FOUR_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("fournisseurs_list"))
    else:
        fd = read_one("FOURNISSEURS",FOUR_COLS,code)
        if not fd: flash("Introuvable.","danger"); return redirect(url_for("fournisseurs_list"))
    return render_template("fournisseurs_form.html", title=f"Modifier fournisseur – {code}",
        action=url_for("fournisseurs_edit",code=code), fd=fd, errors=errors, L=L, is_new=False)

@app.route("/fournisseurs/supprimer/<code>", methods=["POST"])
def fournisseurs_delete(code):
    ok,msg = delete_row("FOURNISSEURS",FOUR_COLS,code)
    flash(msg,"success" if ok else "danger")
    return redirect(url_for("fournisseurs_list"))

# ── PIECES ────────────────────────────────────────────────────────────────────

@app.route("/pieces")
def pieces_list():
    rows = read_sheet("PIECES", PIECE_COLS)
    q  = request.args.get("q","").strip().lower()
    fc = request.args.get("categorie","").strip()
    fa = request.args.get("alerte","").strip()
    if q:       rows = [r for r in rows if any(q in str(v).lower() for v in r.values())]
    if fc:      rows = [r for r in rows if r.get("Categorie") == fc]
    if fa=="1": rows = [r for r in rows if str(r.get("Alerte_Stock","")).startswith("⚠")]
    L = get_listes()
    return render_template("pieces_list.html", rows=rows,
        search=q, f_cat=fc, f_alert=fa, categories=L["Categorie_Piece"])

@app.route("/pieces/nouveau", methods=["GET","POST"])
def pieces_new():
    L = get_listes()
    existing = get_col_values("PIECES",1)
    errors = {}
    fd = {lbl:"" for _,lbl,_ in PIECE_COLS}
    fd["Code_Piece"] = next_code("PC", existing, 3)
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Code_Piece","").strip():  errors["Code_Piece"]="Obligatoire."
        elif fd["Code_Piece"] in existing:       errors["Code_Piece"]="Code déjà utilisé."
        if not fd.get("Designation","").strip(): errors["Designation"]="Obligatoire."
        validate_dates(fd, PIECE_DATES, errors)
        if not errors:
            ok,msg = write_row("PIECES",PIECE_COLS,fd,True,PIECE_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("pieces_list"))
    return render_template("pieces_form.html", title="Nouvelle pièce / consommable",
        action=url_for("pieces_new"), fd=fd, errors=errors, L=L,
        fournisseurs=get_col_values("FOURNISSEURS",1),
        equipements=get_col_values("EQUIPEMENTS",1), is_new=True)

@app.route("/pieces/modifier/<code>", methods=["GET","POST"])
def pieces_edit(code):
    L = get_listes()
    errors = {}
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Designation","").strip(): errors["Designation"]="Obligatoire."
        validate_dates(fd, PIECE_DATES, errors)
        if not errors:
            ok,msg = write_row("PIECES",PIECE_COLS,fd,False,PIECE_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("pieces_list"))
    else:
        fd = read_one("PIECES",PIECE_COLS,code)
        if not fd: flash("Introuvable.","danger"); return redirect(url_for("pieces_list"))
    return render_template("pieces_form.html", title=f"Modifier pièce – {code}",
        action=url_for("pieces_edit",code=code), fd=fd, errors=errors, L=L,
        fournisseurs=get_col_values("FOURNISSEURS",1),
        equipements=get_col_values("EQUIPEMENTS",1), is_new=False)

@app.route("/pieces/supprimer/<code>", methods=["POST"])
def pieces_delete(code):
    ok,msg = delete_row("PIECES",PIECE_COLS,code)
    flash(msg,"success" if ok else "danger")
    return redirect(url_for("pieces_list"))

# ── TECHNICIENS ───────────────────────────────────────────────────────────────

@app.route("/techniciens")
def techniciens_list():
    rows = read_sheet("TECHNICIENS", TECH_COLS)
    q  = request.args.get("q","").strip().lower()
    fs = request.args.get("service","").strip()
    if q:  rows = [r for r in rows if any(q in str(v).lower() for v in r.values())]
    if fs: rows = [r for r in rows if r.get("Service_Rattachement") == fs]
    L = get_listes()
    return render_template("techniciens_list.html", rows=rows,
        search=q, f_serv=fs, services=L["Service"])

@app.route("/techniciens/nouveau", methods=["GET","POST"])
def techniciens_new():
    L = get_listes()
    existing = get_col_values("TECHNICIENS",1)
    errors = {}
    fd = {lbl:"" for _,lbl,_ in TECH_COLS}
    fd["Code_Tech"] = next_code("TECH", existing, 2)
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Code_Tech","").strip(): errors["Code_Tech"]="Obligatoire."
        elif fd["Code_Tech"] in existing:      errors["Code_Tech"]="Code déjà utilisé."
        if not fd.get("Nom","").strip():       errors["Nom"]="Obligatoire."
        if not fd.get("Prenom","").strip():    errors["Prenom"]="Obligatoire."
        if not errors:
            ok,msg = write_row("TECHNICIENS",TECH_COLS,fd,True,TECH_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("techniciens_list"))
    return render_template("techniciens_form.html", title="Nouveau technicien",
        action=url_for("techniciens_new"), fd=fd, errors=errors, L=L, is_new=True)

@app.route("/techniciens/modifier/<code>", methods=["GET","POST"])
def techniciens_edit(code):
    L = get_listes()
    errors = {}
    if request.method == "POST":
        fd = dict(request.form)
        if not fd.get("Nom","").strip():    errors["Nom"]="Obligatoire."
        if not fd.get("Prenom","").strip(): errors["Prenom"]="Obligatoire."
        if not errors:
            ok,msg = write_row("TECHNICIENS",TECH_COLS,fd,False,TECH_FCOLS)
            flash(msg,"success" if ok else "danger")
            if ok: return redirect(url_for("techniciens_list"))
    else:
        fd = read_one("TECHNICIENS",TECH_COLS,code)
        if not fd: flash("Introuvable.","danger"); return redirect(url_for("techniciens_list"))
    return render_template("techniciens_form.html", title=f"Modifier technicien – {code}",
        action=url_for("techniciens_edit",code=code), fd=fd, errors=errors, L=L, is_new=False)

@app.route("/techniciens/supprimer/<code>", methods=["POST"])
def techniciens_delete(code):
    ok,msg = delete_row("TECHNICIENS",TECH_COLS,code)
    flash(msg,"success" if ok else "danger")
    return redirect(url_for("techniciens_list"))

# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    def open_browser():
        import time; time.sleep(1.2)
        webbrowser.open("http://127.0.0.1:5000")
    threading.Thread(target=open_browser, daemon=True).start()
    print("\n" + "="*55)
    print("  GMAO – Interface de saisie locale")
    print("  Ouvrir  : http://127.0.0.1:5000")
    print("  Arrêter : Ctrl+C")
    print("="*55 + "\n")
    app.run(debug=False, port=5000)
